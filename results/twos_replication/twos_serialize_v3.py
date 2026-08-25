#!/usr/bin/env python3
"""TWOS -> serialized JSONL (CERT-analogous DAY/PSY/SES lines).

Grain: --grain tenmin (default) buckets mouse+keystroke behavior into
10-minute windows; hour also supported. PSY = real IPIP-50 OCEAN per user;
DAY = team/machine metadata; SES = per-window behavioral counts. Positive:
a (victim_user, WC-date, window) overlapping a masquerade attack period (the
attacker operates the victim's live session).
"""
import re, glob, os, json, argparse
from collections import defaultdict, Counter
import openpyxl

ROOT = os.path.expanduser("~/twos_raw/TWOS-dataset")

KEY = {}
for i,(tr,sg) in enumerate([
 ('E','+'),('A','-'),('C','+'),('N','+'),('O','+'),('E','-'),('A','+'),('C','-'),('N','-'),('O','-'),
 ('E','+'),('A','-'),('C','+'),('N','-'),('O','+'),('E','-'),('A','+'),('C','-'),('N','-'),('O','-'),
 ('E','+'),('A','-'),('C','+'),('N','+'),('O','+'),('E','-'),('A','+'),('C','-'),('N','+'),('O','-'),
 ('E','+'),('A','-'),('C','+'),('N','+'),('O','+'),('E','-'),('A','+'),('C','-'),('N','+'),('O','+'),
 ('E','+'),('A','+'),('C','+'),('N','+'),('O','+'),('E','-'),('A','+'),('C','+'),('N','+'),('O','+'),
]): KEY[i+1]=(tr,sg)

def score_ocean(path):
    ws = openpyxl.load_workbook(path).active
    traits=defaultdict(list); item=0
    for row in ws.iter_rows(min_row=3, values_only=False):
        if row[0].value is None: continue
        item+=1
        if item>50: break
        xcol=next((j for j,c in enumerate(row) if str(c.value).strip().upper()=="X"), None)
        if xcol is None: continue
        tr,sg=KEY[item]; traits[tr].append(xcol if sg=='+' else 6-xcol)
    return {t: round(sum(v)/len(v),2) for t,v in traits.items() if v}

def load_meta_and_attacks():
    ws=openpyxl.load_workbook(os.path.join(ROOT,"ImportantInfo.xlsx"), data_only=True).active
    rows=[[c for c in r] for r in ws.iter_rows(values_only=True)]
    meta={}
    for r in rows:
        if r and isinstance(r[0],str) and re.fullmatch(r"User\d+", r[0]):
            meta[r[0]]={"team":r[3] or "NA","leader":"yes" if r[2]=="x" else "no","machine":r[6] or "NA","machine2":(r[7] if len(r)>7 else None) and r[8] or None,"team2":r[4] if len(r)>4 else None}
    WC_DATE={"tue":"2017-03-21","thu":"2017-03-23"}
    attacks=[]; day=None
    for r in rows:
        joined=" ".join(str(c) for c in r if c is not None)
        if "WC on Tuesday" in joined: day="tue"
        elif "WC on Thursday" in joined: day="thu"
        elif "Hiring" in joined or "Firing" in joined: day="traitor"
        if day in ("tue","thu") and r and isinstance(r[0],str) and re.fullmatch(r"User\d+", r[0]):
            att=r[11] if len(r)>11 else None; vic=r[12] if len(r)>12 else None
            if att and vic and isinstance(vic,str) and re.fullmatch(r"User\d+", str(vic).strip()):
                m=re.findall(r"(\d{1,2})\.(\d{2})", str(att))
                if len(m)==2:
                    s=int(m[0][0])*60+int(m[0][1]); e=int(m[1][0])*60+int(m[1][1])
                    attacks.append((vic.strip(), WC_DATE[day], s, e))
    return meta, attacks

def user_of(fn):
    m=re.search(r"User(\d+)", os.path.basename(fn)); return "User"+m.group(1) if m else None

def main():
    ap=argparse.ArgumentParser(); ap.add_argument("--grain",choices=["hour","tenmin"],default="tenmin")
    ap.add_argument("--out",default=os.path.expanduser("~/twos_work/session_jsonl")); a=ap.parse_args()
    os.makedirs(a.out, exist_ok=True)
    meta,attacks=load_meta_and_attacks()
    ocean={}
    for f in glob.glob(os.path.join(ROOT,"personality/personality_test_ano/User*.xlsx")):
        try: ocean[user_of(f)]=score_ocean(f)
        except Exception as e: print("ocean fail", f, e)
    TRAITOR_DATE="2017-03-24"; TR_S,TR_E=14*60,16*60
    traitors=sorted([u for u,m in meta.items() if m.get("team2")], key=lambda s:int(s[4:]))
    mach2user={}
    for u,m in meta.items():
        if m.get("machine") and m["machine"]!="NA": mach2user[m["machine"]]=u
        if m.get("machine2"): mach2user[m["machine2"]]=u
    atk=defaultdict(list)
    for vic,d,s,e in attacks: atk[vic].append((d,s,e))
    mal_users=sorted(atk, key=lambda s:int(s[4:]))

    def bkey(u,ts):
        if a.grain=="hour": return (u,ts[:10],ts[11:13]+":00");
        return (u,ts[:10],ts[11:13]+":"+ts[14]+"0")  # HH:M0
    W=defaultdict(lambda: defaultdict(float))
    for fn in glob.glob(os.path.join(ROOT,"mouse/**/*"), recursive=True):
        if not os.path.isfile(fn): continue
        u=user_of(fn)
        if not u: continue
        with open(fn,errors="ignore") as f:
            for line in f:
                if len(line)<15 or not line[:4].isdigit(): continue
                d=W[bkey(u,line)]; ev=line.split(";",2)[1].strip() if ";" in line else ""
                d["mouse_ev"]+=1
                if "Moved" in ev: d["moves"]+=1
                elif "Press" in ev or "Click" in ev: d["clicks"]+=1
    for fn in glob.glob(os.path.join(ROOT,"keystroke/**/*"), recursive=True):
        if not os.path.isfile(fn): continue
        u=user_of(fn)
        if not u: continue
        with open(fn,errors="ignore") as f:
            for line in f:
                s=line.lstrip('"')
                if len(s)<15 or not s[:4].isdigit(): continue
                d=W[bkey(u,s)]; parts=[p.strip('"') for p in line.split(",")]
                d["key_ev"]+=1
                if len(parts)>1 and parts[1]=="press": d["press"]+=1
                if len(parts)>2 and parts[2].startswith("Key."): d["special"]+=1
                if len(parts)>2 and "backspace" in parts[2].lower(): d["backspace"]+=1

    import os as _os
    EVMAP={"LogInSuccess":"ev_ok","LogInAttempt":"ev_att","LogOff":"ev_off"}
    for fn in glob.glob(_os.path.join(ROOT,"eventviewer/**/*.log"), recursive=True):
        mach=_os.path.basename(fn)[:-4]; u=mach2user.get(mach)
        if not u: continue
        with open(fn,errors="ignore") as f:
            for line in f:
                parts=[x.strip(chr(34)) for x in line.strip().split(",")]
                if len(parts)<3 or "/" not in parts[0]: continue
                try:
                    mm,dd,yy=parts[0].split(" ")[0].split("/"); hhmm=parts[0].split(" ")[1][:5]
                    date="20"+yy+"-"+mm+"-"+dd
                except Exception: continue
                win=hhmm[:2]+":"+hhmm[3]+"0" if True else hhmm
                ev=EVMAP.get(parts[2])
                if ev: W[(u,date,win)][ev]+=1
    def is_pos(u,date,win):
        hh,mm=int(win[:2]),int(win[3:]); t=hh*60+mm
        for d,s,e in atk.get(u,[]):
            if d==date and s<=t<e: return (1,"masq")
        if u in traitors and date==TRAITOR_DATE and TR_S<=t<TR_E: return (1,"traitor")
        return (0,"none")

    days=sorted({(u,dt) for (u,dt,_) in W}); di={ud:i for i,ud in enumerate(days)}
    rows=[]
    for (u,dt,win),feat in sorted(W.items()):
        oc=ocean.get(u,{}); psy=" ".join(f"{t}={oc.get(t,'NA')}" for t in ["O","C","E","A","N"])
        m=meta.get(u,{})
        text="\n".join([f"DAY team={m.get('team','NA')} leader={m.get('leader','no')} machine={m.get('machine','NA')}",
            "PSY "+psy, "SESSIONS total=1 kept=1",
            "SES win={} mouse_ev={:.0f} moves={:.0f} clicks={:.0f} key_ev={:.0f} press={:.0f} special={:.0f} backspace={:.0f} ev_ok={:.0f} ev_att={:.0f} ev_off={:.0f}".format(
              win,feat['mouse_ev'],feat['moves'],feat['clicks'],feat['key_ev'],feat['press'],feat['special'],feat['backspace'],feat['ev_ok'],feat['ev_att'],feat['ev_off'])])
        _y,_at=is_pos(u,dt,win)
        rows.append({"example_id":f"{u}:{dt}:{win}","user_id":u,"day_index":di[(u,dt)],
                     "date":dt,"win":win,"y":_y,"atype":_at,
                     "n_sessions_total":1,"n_sessions_kept":1,
                     "context":{"team":meta.get(u,{}).get("team","NA")},"text":text})
    posusers=set(mal_users)
    for r in rows:
        if r["user_id"] in posusers: r["split"]="eval"
        elif r["user_id"] in traitors: r["split"]="eval" if r["date"]==TRAITOR_DATE else ("val" if hash(r["example_id"])%10==0 else "train")
        else: r["split"]="val" if hash(r["example_id"])%10==0 else "train"
    fhs={s:open(os.path.join(a.out,s+".jsonl"),"w") for s in ("eval","train","val")}
    for r in rows: fhs[r["split"]].write(json.dumps(r)+"\n")
    for fh in fhs.values(): fh.close()
    npos=sum(r["y"] for r in rows); sc=Counter(r["split"] for r in rows)
    print(f"[grain={a.grain}] OCEAN users={len(ocean)}; masq attacks={len(attacks)}; malicious users={len(mal_users)}")
    print(f"total windows={len(rows)} positives={npos} ({100*npos/max(len(rows),1):.1f}%)  split={dict(sc)}")
    print(f"benign train pool users={len(set(r['user_id'] for r in rows if r['split']!='eval'))}")
    from collections import Counter as _C
    print("positives by type:", dict(_C(r['atype'] for r in rows if r['y']==1)))
    print("traitors:", traitors)

if __name__=="__main__": main()
